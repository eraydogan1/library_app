import os, sys, tkinter as tk
from tkinter import messagebox, filedialog
from tkinter.ttk import Notebook, Treeview, Combobox, Style, Separator
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from PIL import Image, ImageTk 


def base():
    if getattr(sys, "frozen", False): return os.path.dirname(sys.executable)
    return os.path.dirname(__file__)
BASE = base()
def p(x): return os.path.join(BASE, x)


def get_next_id(file_name):
    try:
        if not os.path.exists(p(file_name)): return 1
        wb = load_workbook(p(file_name)); ws = wb.active
        ids = [int(row[0]) for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0] is not None and isinstance(row[0], int)]
        return max(ids) + 1 if ids else 1
    except: return 1


def ensure():
    files = {
        "ogrenciler.xlsx": ["Barkod", "No", "Ad", "Sınıf"],
        "kitaplar.xlsx":   ["Barkod", "ISBN", "Ad", "Yazar"], 
        "odunc.xlsx":      ["IslemID", "KitapBarkod", "OgrBarkod", "Alis", "SonGun", "Durum"],
        "teslim.xlsx":     ["IslemID", "KitapBarkod", "OgrBarkod", "Alis", "Teslim"],
        "mezunlar.xlsx":   ["Barkod", "No", "Ad", "MezuniyetYili"]
    }
    for f, h in files.items():
        if not os.path.exists(p(f)):
            wb = Workbook(); ws = wb.active; ws.append(h); wb.save(p(f))
ensure()


root = tk.Tk()
root.title("📚 Tam Barkodlu Kütüphane Sistemi (V2 - Resimli)")
root.state("zoomed")
root.configure(bg="#f4f6f7")


global_bg_image = None

def arka_plan_yukle():
    global global_bg_image
    resim_yolu = p("arkaplan.png")
    
    if os.path.exists(resim_yolu):
        try:
        
            sw = root.winfo_screenwidth()
            sh = root.winfo_screenheight()
            
          
            img = Image.open(resim_yolu)
            img = img.resize((sw, sh), Image.Resampling.LANCZOS)
            global_bg_image = ImageTk.PhotoImage(img)
            return True
        except Exception as e:
            print(f"Resim hatası: {e}")
            return False
    return False


resim_var_mi = arka_plan_yukle()

def resim_koy(frame):
    """Verilen Frame'in arkasına resmi döşer"""
    if resim_var_mi and global_bg_image:
        bg_lbl = tk.Label(frame, image=global_bg_image)
        bg_lbl.place(x=0, y=0, relwidth=1, relheight=1)

style = Style()
style.theme_use("default")
style.configure("Treeview", background="white", foreground="#2c3e50", rowheight=28, fieldbackground="white", font=("Arial", 11))
style.configure("Treeview.Heading", font=("Arial", 12, "bold"), background="#d6dbdf", foreground="#2c3e50")
style.map("Treeview", background=[("selected", "#aed6f1")])

nb = Notebook(root)
nb.pack(fill="both", expand=True, padx=8, pady=8)

tab_main = tk.Frame(nb, bg="#f4f6f7")
tab_add  = tk.Frame(nb, bg="#f4f6f7") 
tab_io   = tk.Frame(nb, bg="#f4f6f7") 
tab_reports = tk.Frame(nb, bg="#f4f6f7") 
tab_grads = tk.Frame(nb, bg="#f4f6f7")
tab_search = tk.Frame(nb, bg="#f4f6f7")


resim_koy(tab_add) 
resim_koy(tab_io)


nb.add(tab_main, text="📊 Ana Tablolar")
nb.add(tab_add,  text="➕ Ekle (Barkodlu)")
nb.add(tab_io,   text="🔄 Ödünç / Teslim")
nb.add(tab_reports, text="📈 Raporlar")
nb.add(tab_grads, text="🎓 Mezunlar") 
nb.add(tab_search, text="🔍 Arama / Silme") 


def table(frame, cols, title):
    f_c = tk.Frame(frame, bg="#f4f6f7"); f_c.pack(fill="both", expand=True, padx=4, pady=4)
    tk.Label(f_c, text=title, font=("Arial",16,"bold"), bg="#f4f6f7", fg="#2c3e50").pack(pady=6)
    tv = Treeview(f_c, columns=cols, show="headings")
    vsb = tk.Scrollbar(f_c, orient="vertical", command=tv.yview); hsb = tk.Scrollbar(f_c, orient="horizontal", command=tv.xview)
    tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x"); tv.pack(side="left", fill="both", expand=True)
    for c in cols: tv.heading(c, text=c); tv.column(c, width=120, anchor="center")
    
    tv.tag_configure('koyu_kirmizi', background='#e6b0aa') 
    tv.tag_configure('acik_kirmizi', background='#f2d7d5') 
    tv.tag_configure('sari', background='#fcf3cf')         
    tv.tag_configure('normal', background='white')         
    return tv


top = tk.Frame(tab_main, bg="#f4f6f7"); top.pack(fill="both", expand=True)
bot = tk.Frame(tab_main, bg="#f4f6f7"); bot.pack(fill="both", expand=True)
tv_students = table(top, ["Barkod","No","Ad","Sınıf"], "👩‍🎓 AKTİF ÖĞRENCİLER")
tv_books    = table(top, ["Barkod","ISBN","Ad","Yazar"], "📘 KİTAP ENVANTERİ") 


report_f_frame = tk.Frame(tab_reports, bg="#f4f6f7")
report_f_frame.pack(fill="x", padx=20, pady=10)
tk.Label(report_f_frame, text="Durum Filtresi:", font=("Arial", 11, "bold"), bg="#f4f6f7").pack(side="left", padx=5)
combo_filter = Combobox(report_f_frame, values=["Hepsi", "Koyu Kırmızı (2 Gün)", "Açık Kırmızı (5 Gün)", "Sarı (8 Gün)"], state="readonly", width=25)
combo_filter.current(0)
combo_filter.pack(side="left", padx=5)

tv_loans_out = table(tab_reports, ["IslemID","KitapBarkod","Kitap Adı","OgrBarkod","Öğrenci Adı","Alış Tarihi","Son Teslim"], "📕 ÖDÜNÇ LİSTESİ")
tv_performance = table(tab_reports, ["Barkod","Öğrenci Adı","Durum"], "📈 ÖĞRENCİ LİSTESİ")

tv_loans = table(bot, ["IslemID","KitapBarkod","OgrBarkod","Alis","SonGun","Durum"], "TÜM KAYITLAR")
tv_returns = table(bot, ["IslemID","KitapBarkod","OgrBarkod","Alis","Teslim"], "📗 TESLİM GEÇMİŞİ")
tv_mezunlar = table(tab_grads, ["Barkod", "No", "Ad", "MezuniyetYili"], "🎓 MEZUN ARŞİVİ")


def parse_date_safe(val):
    if isinstance(val, datetime): return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if not val: return None
    v_s = str(val).split(' ')[0]
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try: return datetime.strptime(v_s, fmt)
        except: continue
    return None

def refresh(search_term=""):
    for tv in (tv_students, tv_books, tv_loans, tv_returns, tv_loans_out, tv_performance, tv_mezunlar):
        for i in tv.get_children(): tv.delete(i)
    
    b_c, s_c = {}, {} 
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    sel_filter = combo_filter.get()


    wb_o = load_workbook(p("ogrenciler.xlsx")); ws_o = wb_o.active
    for r in ws_o.iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        tv_students.insert("", "end", values=r)
        s_c[str(r[0])] = str(r[2]) 
    
    
    wb_m = load_workbook(p("mezunlar.xlsx")); ws_m = wb_m.active
    for r in ws_m.iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        tv_mezunlar.insert("", "end", values=r)
        s_c[str(r[0])] = str(r[2]) + " (Mezun)"


    wb_k = load_workbook(p("kitaplar.xlsx")); ws_k = wb_k.active
    for r in ws_k.iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        tv_books.insert("", "end", values=r)
        b_c[str(r[0])] = str(r[2]) 

   
    wb_od = load_workbook(p("odunc.xlsx")); ws_od = wb_od.active
    for r in ws_od.iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        durum = r[5] if r[5] else "Ödünç"
        tv_loans.insert("", "end", values=r)
        
        if durum == "Ödünç":
            k_bar = str(r[1])
            o_bar = str(r[2])
            ka = b_c.get(k_bar, "Silinmiş Kitap")
            oa = s_c.get(o_bar, "Bilinmeyen Kişi")
            
            tag = 'normal'; son_gun = parse_date_safe(r[4])
            if son_gun:
                fark = (son_gun - today).days
                if fark <= 2: tag = 'koyu_kirmizi'
                elif fark <= 5: tag = 'acik_kirmizi'
                elif fark <= 8: tag = 'sari'

            show = False
            if sel_filter == "Hepsi": show = True
            elif sel_filter == "Koyu Kırmızı (<=2 Gün)" and tag == 'koyu_kirmizi': show = True
            elif sel_filter == "Açık Kırmızı (<=5 Gün)" and tag == 'acik_kirmizi': show = True
            elif sel_filter == "Sarı (<=8 Gün)" and tag == 'sari': show = True

            if show and (search_term == "" or search_term.lower() in f"{k_bar} {ka} {oa} {o_bar}".lower()):
                tv_loans_out.insert("", "end", values=(r[0],k_bar,ka,o_bar,oa,r[3],r[4]), tags=(tag,))

combo_filter.bind("<<ComboboxSelected>>", lambda e: refresh())


def ogr_ekle(event=None):
    bar = e_bar.get().strip()
    if not bar: return
    
    wb = load_workbook(p("ogrenciler.xlsx")); ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == bar:
            messagebox.showerror("Hata", "Bu barkod zaten kayıtlı!"); return

    ws.append([bar, e_no.get(), e_ad.get(), e_cls.get()])
    wb.save(p("ogrenciler.xlsx"))
    refresh()
    e_bar.delete(0, 'end'); e_no.delete(0, 'end'); e_ad.delete(0, 'end')
    e_bar.focus_set() 

def kitap_ekle(event=None):
    isbn = k_isbn.get().strip()
    bar  = k_bar.get().strip()
    if not bar: 
        messagebox.showwarning("Uyarı", "Kütüphane Barkodu boş olamaz!")
        return

    wb = load_workbook(p("kitaplar.xlsx")); ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == bar:
            messagebox.showerror("Hata", "Bu kütüphane barkodu zaten var!"); return

    ws.append([bar, isbn, k_ad.get(), k_y.get()])
    wb.save(p("kitaplar.xlsx"))
    refresh()
    k_isbn.delete(0, 'end'); k_bar.delete(0, 'end'); k_ad.delete(0, 'end'); k_y.delete(0, 'end')
    k_isbn.focus_set()


def odunc(event=None):
    kbar, obar = l_book.get().strip(), l_no.get().strip()
    if not kbar or not obar: return
    
    wb = load_workbook(p("odunc.xlsx")); ws = wb.active
    for r in ws.iter_rows(min_row=2, values_only=True):
        if str(r[1]) == kbar and r[5] == "Ödünç":
             messagebox.showerror("Hata", "Bu kitap şu an başkasında!"); return

    nid = get_next_id("odunc.xlsx")
    a, s = datetime.now(), datetime.now() + timedelta(days=15)
    ws.append([nid, kbar, obar, a.strftime("%d.%m.%Y"), s.strftime("%d.%m.%Y"), "Ödünç"])
    wb.save(p("odunc.xlsx"))
    refresh()
    l_book.delete(0, 'end'); l_no.delete(0, 'end')
    l_book.focus_set()

def teslim(event=None):
    kbar, obar = l_book.get().strip(), l_no.get().strip()
    wb_o = load_workbook(p("odunc.xlsx")); ws_o = wb_o.active
    found = False
    
    for r in ws_o.iter_rows(min_row=2):
        if str(r[1].value)==kbar and str(r[2].value)==obar and r[5].value=="Ödünç":
            r[5].value="Teslim"
            wb_t = load_workbook(p("teslim.xlsx")); ws_t = wb_t.active
            ws_t.append([get_next_id("teslim.xlsx"), kbar, obar, r[3].value, datetime.now().strftime("%d.%m.%Y")])
            wb_t.save(p("teslim.xlsx"))
            found = True
            break
    
    if found:
        wb_o.save(p("odunc.xlsx")); refresh()
        l_book.delete(0, 'end'); l_no.delete(0, 'end')
        messagebox.showinfo("Başarılı", "Teslim alındı.")
        l_book.focus_set()
    else:
        messagebox.showerror("Hata", "Eşleşen aktif ödünç kaydı bulunamadı!")


def toplu_sinif_atlat():
    if not messagebox.askyesno("Yıl Sonu", "Tüm sınıflar bir üst seviyeye taşınacak. 4. sınıflar MEZUN olacak. Emin misin?"): return
    wb_o = load_workbook(p("ogrenciler.xlsx")); ws_o = wb_o.active
    wb_m = load_workbook(p("mezunlar.xlsx")); ws_m = wb_m.active
    silinecek, m_say, a_say, yil = [], 0, 0, datetime.now().year

    for row in ws_o.iter_rows(min_row=2):
        if row[0].value is None: continue
        sinif_cell = row[3]
        sinif_txt = str(sinif_cell.value).strip()
        
        if "-" in sinif_txt:
            try:
                parts = sinif_txt.split("-"); s_no = int(parts[0]); sube = parts[1]
                if s_no < 4:
                    sinif_cell.value = f"{s_no + 1}-{sube}"; a_say += 1
                else:
                    ws_m.append([row[0].value, row[1].value, row[2].value, f"{yil} Mezunu"])
                    silinecek.append(row[0].row); m_say += 1
            except: continue

    for r_idx in reversed(silinecek): ws_o.delete_rows(r_idx)
    wb_o.save(p("ogrenciler.xlsx")); wb_m.save(p("mezunlar.xlsx")); refresh()
    messagebox.showinfo("Tamam", f"{a_say} sınıf atladı, {m_say} mezun oldu.")

def delete_item():
    sel_s, sel_k, sel_m = tv_students.selection(), tv_books.selection(), tv_mezunlar.selection()
    if sel_s: f, val, col = "ogrenciler.xlsx", tv_students.item(sel_s,'values')[0], 0
    elif sel_k: f, val, col = "kitaplar.xlsx", tv_books.item(sel_k,'values')[0], 0 
    elif sel_m: f, val, col = "mezunlar.xlsx", tv_mezunlar.item(sel_m,'values')[0], 0
    else: return
    if messagebox.askyesno("Sil", "Seçili kayıt silinsin mi?"):
        wb = load_workbook(p(f)); ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[col].value) == str(val): ws.delete_rows(row[0].row); break
        wb.save(p(f)); refresh()


def lbl(parent, text, color=None):
    
    bg_c = color if color else "#f4f6f7"
    return tk.Label(parent, text=text, bg=bg_c, font=("Arial", 10, "bold"))


tk.Label(tab_add, text="👩‍🎓 ÖĞRENCİ EKLE", font=("Arial", 14, "bold"), bg="#ffffff", relief="solid").grid(row=0, column=0, pady=20, padx=10, sticky="w")

lbl(tab_add, "Öğrenci Barkod:", "#ffffff").grid(row=1, column=0, sticky="w", padx=10); 
e_bar = tk.Entry(tab_add); e_bar.grid(row=1, column=1, pady=5)

lbl(tab_add, "Okul No:", "#ffffff").grid(row=2, column=0, sticky="w", padx=10); 
e_no = tk.Entry(tab_add); e_no.grid(row=2, column=1, pady=5)

lbl(tab_add, "Ad Soyad:", "#ffffff").grid(row=3, column=0, sticky="w", padx=10); 
e_ad = tk.Entry(tab_add); e_ad.grid(row=3, column=1, pady=5)

e_cls = Combobox(tab_add, values=["1-A","2-A","3-A","4-A"]); e_cls.grid(row=4, column=1, pady=5); e_cls.current(0)
tk.Button(tab_add, text="Öğrenci Ekle", command=ogr_ekle, bg="#82e0aa").grid(row=5, column=1, pady=10)
e_bar.bind('<Return>', lambda event: e_no.focus_set()) 


tk.Label(tab_add, text="📘 KİTAP EKLE", font=("Arial", 14, "bold"), bg="#ffffff", relief="solid").grid(row=6, column=0, pady=20, padx=10, sticky="w")

lbl(tab_add, "ISBN / Orijinal:", "#ffffff").grid(row=7, column=0, sticky="w", padx=10)
k_isbn = tk.Entry(tab_add); k_isbn.grid(row=7, column=1, pady=5)

lbl(tab_add, "Kütüphane Etiketi:", "#ffffff").grid(row=8, column=0, sticky="w", padx=10)
k_bar = tk.Entry(tab_add); k_bar.grid(row=8, column=1, pady=5)

lbl(tab_add, "Kitap Adı:", "#ffffff").grid(row=9, column=0, sticky="w", padx=10)
k_ad = tk.Entry(tab_add); k_ad.grid(row=9, column=1, pady=5)

lbl(tab_add, "Yazar:", "#ffffff").grid(row=10, column=0, sticky="w", padx=10)
k_y = tk.Entry(tab_add); k_y.grid(row=10, column=1, pady=5)

tk.Button(tab_add, text="Kitap Ekle", command=kitap_ekle, bg="#85c1e9").grid(row=11, column=1, pady=10)

k_isbn.bind('<Return>', lambda event: k_bar.focus_set())
k_bar.bind('<Return>', lambda event: k_ad.focus_set())
k_y.bind('<Return>', kitap_ekle)


tk.Frame(tab_add, bg="black", height=2).grid(row=12, column=0, columnspan=2, sticky="ew", pady=20)
tk.Button(tab_add, text="🎓 Yıl Sonu (Sınıf Atlat)", command=toplu_sinif_atlat, bg="#d98880", fg="white").grid(row=13, column=0, columnspan=2)


center_io = tk.Frame(tab_io, bg="#ffffff", bd=5, relief="ridge")
center_io.pack(pady=100)

tk.Label(center_io, text="Kütüphane Etiketi (Barkod):", font=("Arial",12), bg="#ffffff").pack(pady=5)
l_book = tk.Entry(center_io, font=("Arial", 14)); l_book.pack(pady=5, padx=20)

tk.Label(center_io, text="Öğrenci Barkod:", font=("Arial",12), bg="#ffffff").pack(pady=5)
l_no = tk.Entry(center_io, font=("Arial", 14)); l_no.pack(pady=5, padx=20)

l_book.bind('<Return>', lambda event: l_no.focus_set()) 
l_no.bind('<Return>', odunc)

s_f = tk.Frame(tab_search, bg="#f4f6f7"); s_f.pack(pady=20)
e_search = tk.Entry(s_f, width=40); e_search.pack(pady=5)
tk.Button(s_f, text="🔍 Ara", command=lambda: refresh(e_search.get())).pack()
tk.Button(tab_search, text="🗑️ Sil (Seçili Olanı)", bg="#c0392b", fg="white", command=delete_item).pack(pady=20)
tk.Button(tab_main, text="🔄 Verileri Yenile", command=refresh).pack(pady=10)

root.after(100, lambda: refresh())
root.mainloop()